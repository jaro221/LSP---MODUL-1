# -*- coding: utf-8 -*-
"""
Created on Mon Mar 10 11:08:20 2025

@author: jarom
"""

import os

import tensorflow as tf
import numpy as np
from keras.layers import Conv2D, Input, BatchNormalization, LeakyReLU, ZeroPadding2D, MaxPool2D, Activation
from keras import layers
from keras.layers import add, concatenate
from keras.models import Model
from keras import losses

from tensorflow.keras.callbacks import ModelCheckpoint,EarlyStopping

from platform import python_version
import json
import pandas as pd
import openpyxl
import random

print("Python version: {}".format(python_version()))
print("TensorFlow version: {}".format(tf.__version__))
print("Eager execution: {}".format(tf.executing_eagerly()))
print("Keras version: {}".format(tf.keras.__version__))

print("Num Physical GPUs Available: ", len(tf.config.experimental.list_physical_devices('GPU')))
print("Num Logical GPUs Available: ", len(tf.config.experimental.list_logical_devices('GPU')))

NETWORK_W          = 332
NETWORK_H          = 332
name="model_modul1_v5_Large_GLOB"
path="D:/Projekty/2022_01_BattPor/DATA_M1_v5/"

class Data_Generator(tf.keras.utils.Sequence):  
    def __init__(self, batch_size, root_path, IMG_SIZE, type_gen="Train", grouping=True, split_ratio=0.8, input_dirs=None, label_dirs=None, use_cached_extrems=True, run_extrem=True, verbose=0): 
        self.root_path = root_path
        self.input_dirs = input_dirs if input_dirs else ['']
        self.label_dirs = label_dirs if label_dirs else ['']
        self.type_gen = type_gen
        self.shuffle = True
        self.grouping = grouping
        self.split_ratio = split_ratio
        self.img_size = IMG_SIZE
        self.batch_size = batch_size
        self.use_cached_extrems = use_cached_extrems
        self.run_extrem = run_extrem
        self.verbose=verbose

        self.extrem_values = self.get_extrem() if run_extrem else None
        data_x, data_y = self.get_data()


    def __len__(self):
        return (np.ceil(len(self.data_x) / float(self.batch_size))).astype(np.int32)

    def __getitem__(self, idx):
        try:
            train_x, train_y = self.build_train_generator(idx)
            return (
                {f"input_{i+1}": train_x[i] for i in range(8)},
                {"output_1": train_y[0, :], "output_2": train_y[1, :]}
            )
        except Exception as e:
            if self.verbose!=0:
                print(f"Skipping batch {idx} due to error: {e}")
            return self.__getitem__((idx + 1) % len(self))

    def get_extrem(self):
        cache_file = os.path.join(self.root_path, "extrems.json")
        if self.use_cached_extrems and os.path.exists(cache_file):
            if self.verbose!=0:
                print("Loading cached extrem values from extrems.json")
            with open(cache_file, 'r') as f:
                extrem_data = json.load(f)
                extrem_values = np.array([[v["Min"], v["Max"]] for v in extrem_data.values()])
                return extrem_values

        indexes = [1, 2, 3, 4, 7, 8, 9, 10]
        extrem_values = np.full((10, 2), [1e10, -1e10], dtype="float")

        for input_dir, label_dir in zip(self.input_dirs, self.label_dirs):
            input_path = os.path.join(self.root_path, input_dir)
            label_path = os.path.join(self.root_path, label_dir)

            for file_name in os.listdir(input_path):
                try:
                    datax_idx = pd.read_json(os.path.join(input_path, file_name))
                    itemy = file_name[:-5] + ".txt"
                    with open(os.path.join(label_path, itemy), "r") as f:
                        datay_idx = np.fromstring(f.read(), dtype=np.float32, sep=' ')

                    datax_idx_val = datax_idx.values
                    strh = int(datax_idx_val[0, 5])

                    for ddx, cdx in enumerate(indexes):
                        data_cdx = np.asarray(datax_idx_val[0, cdx])
                        if ddx < 4:
                            slice_data = data_cdx[strh-20:strh+380]
                            extrem_values[ddx, 0] = min(extrem_values[ddx, 0], np.min(slice_data))
                            extrem_values[ddx, 1] = max(extrem_values[ddx, 1], np.max(slice_data))
                        else:
                            non_zero = data_cdx[data_cdx > 0]
                            if non_zero.size > 0:
                                extrem_values[ddx, 0] = min(extrem_values[ddx, 0], np.min(non_zero))
                                extrem_values[ddx, 1] = max(extrem_values[ddx, 1], np.max(non_zero))

                    for cdx in range(2):
                        extrem_values[cdx+8, 0] = min(extrem_values[cdx+8, 0], datay_idx[cdx])
                        extrem_values[cdx+8, 1] = max(extrem_values[cdx+8, 1], datay_idx[cdx])

                except Exception as e:
                    if self.verbose!=0:
                        print(f"Extrem error in {file_name}: {e}")
                    continue

        # Save to cache
        extrem_dict = {str(i): {"Min": float(extrem_values[i][0]), "Max": float(extrem_values[i][1])} for i in range(10)}
        with open(cache_file, 'w') as f:
            json.dump(extrem_dict, f)

        return extrem_values

    def get_data(self):
        self.all_data = []
        for input_dir, label_dir in zip(self.input_dirs, self.label_dirs):
            input_path = os.path.join(self.root_path, input_dir)
            label_path = os.path.join(self.root_path, label_dir)
            if not os.path.exists(input_path) or not os.path.exists(label_path):
                continue
            for file in os.listdir(input_path):
                if file.endswith('.json'):
                    self.all_data.append((os.path.join(self.root_path,input_dir, file), os.path.join(self.root_path,label_dir, file[:-5] + ".txt")))

        if self.shuffle:
            random.shuffle(self.all_data)
        data_x, data_y = self.init_data()
        return list(data_x), list(data_y)
            
    def init_data(self):
        split_idx = int(len(self.all_data) * self.split_ratio)
        if self.type_gen == "Train":
            selected = self.all_data[:split_idx]
        else:
            selected = self.all_data[split_idx:]

        data_x, data_y = zip(*selected)
        stop_batch = (len(data_x) // self.batch_size) * self.batch_size
        self.data_x = data_x[:stop_batch]
        self.data_y = data_y[:stop_batch]
        return list(data_x), list(data_y)

    def build_train_generator(self, idx):
        indexes = [1, 2, 3, 4, 7, 8, 9, 10]
        train_x = []
        train_y = []
        train_x_pre = np.zeros((self.batch_size, 8), dtype="object")
        stop = np.zeros((self.batch_size, 2), dtype="float32")
        train_y_pre = np.zeros((self.batch_size, 2), dtype="float32")

        for cdx in range(self.batch_size):
            itemx_path = os.path.join(self.data_x[self.batch_size * idx + cdx])
            itemy_path = os.path.join(self.data_y[self.batch_size * idx + cdx])

            datax_idx = pd.read_json(itemx_path)
            with open(itemy_path, "r") as f:
                datay_idx = np.fromstring(f.read(), dtype=np.float32, sep=' ')
            datax_idx_val = datax_idx.values
            strh = int(datax_idx_val[0, 5])
            stop[cdx, 0] = strh

            for indx, ind in enumerate(indexes):
                train_x_pre[cdx, indx] = datax_idx_val[0, ind]
            train_y_pre[cdx, :] = datay_idx

        for indx, ind in enumerate(indexes):
            if indx < 4:
                train_x_post = np.zeros((self.batch_size, 400), dtype="float")
                for idxx in range(self.batch_size):
                    norm_data = train_x_pre[idxx, indx][int(stop[idxx, 0])-20:int(stop[idxx, 0])+380]
                    train_x_post[idxx, :] = norm_data / self.extrem_values[indx, 1]
            else:
                train_x_post = np.zeros((self.batch_size, self.img_size[0], self.img_size[1]), dtype="float")
                for idxx in range(self.batch_size):
                    train_x_post[idxx, :, :] = train_x_pre[idxx, indx] / self.extrem_values[indx, 1]
            train_x.append(train_x_post)

        train_y = np.asarray([train_y_pre[:, 0], train_y_pre[:, 1]])
        return train_x, train_y




def _conv_block(inp, convs, skip=False):
    x = inp
    count = 0
    
    for conv in convs:
        if count == (len(convs) - 2) and skip:
            skip_connection = x
        count += 1
        
        if conv['stride'] > 1: x = ZeroPadding2D(((1,0),(1,0)), name='zerop_' + str(conv['layer_idx']))(x)  # peculiar padding as darknet prefer left and top
        
        x = Conv2D(conv['filter'], conv['kernel'], strides=conv['stride'], 
                   padding='valid' if conv['stride'] > 1 else 'same', # peculiar padding as darknet prefer left and top
                   name='convn_' + str(conv['layer_idx']) if conv['bnorm'] else 'conv_' + str(conv['layer_idx']),
                   use_bias=True)(x)
        
        if conv['bnorm']: x = BatchNormalization(name='BN_' + str(conv['layer_idx']))(x)    
        
        if conv['activ'] == 1: x = LeakyReLU(alpha=0.1, name='leaky_' + str(conv['layer_idx']))(x)
        if conv['activ'] == 2: x = Activation('mish', name='mish_' + str(conv['layer_idx']))(x) 
            
    return add([skip_connection, x],  name='add_' + str(conv['layer_idx']+1)) if skip else x



if False:
    """ old for v1-v6"""
    I5 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_5')
    
    x = _conv_block(I5, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 0}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_109')(x) 
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 16}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_110')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 23}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_111')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 47}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_147')(x) 
    x5 = _conv_block(x, [{'filter': 24, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 3}])
    
    I6 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_6')
    
    x = _conv_block(I6, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 4}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_112')(x)  
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 17}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_113')(x)  
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 22}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_114')(x)  
    x = _conv_block(x,  [{'filter': 64, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 42}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_144')(x) 
    x6 = _conv_block(x, [{'filter': 24, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 7}])
    
    I7 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_7')
    
    x = _conv_block(I7, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 8}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_115')(x) 
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 18}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_116')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 26}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_117')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 45}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_145')(x) 
    x7 = _conv_block(x, [{'filter': 24, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 11}])
    
    I8 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_8')
    
    x = _conv_block(I8, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 12}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_118')(x) 
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 19}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_119')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 20}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_120')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 46}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_146')(x) 
    x8 = _conv_block(x, [{'filter': 24, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 15}])
    
    I1 = Input(shape=(400,), name="input_1")
    x = layers.Dense(400, activation="relu", use_bias=True)(I1)
    x1 = layers.Dense(600, activation="relu", use_bias=True)(x)
    
    I2 = Input(shape=(400,), name='input_2')
    x = layers.Dense(400, activation="relu", use_bias=True)(I2)
    x2 = layers.Dense(600, activation="relu", use_bias=True)(x)
    
    I3 = Input(shape=(400,), name='input_3')
    x = layers.Dense(400, activation="relu", use_bias=True)(I3)
    x3 = layers.Dense(600, activation="relu", use_bias=True)(x)
    
    I4 = Input(shape=(400,), name='input_4')
    x = layers.Dense(400, activation="relu", use_bias=True)(I4)
    x4 = layers.Dense(600, activation="relu", use_bias=True)(x)
    
    xc5=layers.Flatten()(x5)
    xc6=layers.Flatten()(x6)
    xc7=layers.Flatten()(x7)
    xc8=layers.Flatten()(x8)
    
    xc = concatenate([x1,x2,x3,x4,xc5,xc6,xc7,xc8],  name='concat_1')
    
    xo1 = layers.Dense(4096, activation="relu", use_bias=True)(xc)
    xo1 = layers.Dense(4096, activation="relu", use_bias=True)(xo1)
    xo1 = layers.Dense(1024, activation="relu", use_bias=True)(xo1)
    xo1 = layers.Dense(512, activation="relu", use_bias=True)(xo1)
    xo1 = layers.Dense(1,activation='linear',name="output_1")(xo1)
    model_out1 = xo1
         
    xo2 = layers.Dense(4096, activation="relu", use_bias=True)(xc)
    xo2 = layers.Dense(1024, activation="relu", use_bias=True)(xo2)
    xo2 = layers.Dense(512, activation="relu", use_bias=True)(xo2)
    xo2 = layers.Dense(1,activation='linear',name="output_2")(xo2)
    model_out2 = xo2


if False:
    """ v7-"""
    I5 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_5')
    
    x = _conv_block(I5, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 0}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_109')(x) 
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 16}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_110')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 23}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_111')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 47}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_147')(x) 
    x5 = _conv_block(x, [{'filter': 24, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 3}])
    
    I6 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_6')
    
    x = _conv_block(I6, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 4}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_112')(x)  
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 17}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_113')(x)  
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 22}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_114')(x)  
    x = _conv_block(x,  [{'filter': 64, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 42}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_144')(x) 
    x6 = _conv_block(x, [{'filter': 24, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 7}])
    
    I7 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_7')
    
    x = _conv_block(I7, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 8}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_115')(x) 
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 18}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_116')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 26}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_117')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 45}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_145')(x) 
    x7 = _conv_block(x, [{'filter': 24, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 11}])
    
    I8 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_8')
    
    x = _conv_block(I8, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 12}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_118')(x) 
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 19}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_119')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 20}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_120')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 46}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_146')(x) 
    x8 = _conv_block(x, [{'filter': 24, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 15}])
    
    I1 = Input(shape=(400,), name="input_1")
    x = layers.Dense(400, activation="relu", use_bias=True)(I1)
    x1 = layers.Dense(600, activation="relu", use_bias=True)(x)
    
    I2 = Input(shape=(400,), name='input_2')
    x = layers.Dense(400, activation="relu", use_bias=True)(I2)
    x2 = layers.Dense(600, activation="relu", use_bias=True)(x)
    
    I3 = Input(shape=(400,), name='input_3')
    x = layers.Dense(400, activation="relu", use_bias=True)(I3)
    x3 = layers.Dense(600, activation="relu", use_bias=True)(x)
    
    I4 = Input(shape=(400,), name='input_4')
    x = layers.Dense(400, activation="relu", use_bias=True)(I4)
    x4 = layers.Dense(600, activation="relu", use_bias=True)(x)
    
    xc5=layers.Flatten()(x5)
    xc6=layers.Flatten()(x6)
    xc7=layers.Flatten()(x7)
    xc8=layers.Flatten()(x8)
    
    xc = concatenate([x1,x2,x3,x4,xc5,xc6,xc7,xc8],  name='concat_1')
    xc = layers.Dense(8192, activation="relu", use_bias=True)(xc)
    
    xo1 = layers.Dense(4096, activation="relu", use_bias=True)(xc)
    xo1 = layers.Dense(4096, activation="relu", use_bias=True)(xo1)
    xo1 = layers.Dense(1024, activation="relu", use_bias=True)(xo1)
    xo1 = layers.Dense(512, activation="relu", use_bias=True)(xo1)
    xo1 = layers.Dense(1,activation='linear',name="output_1")(xo1)
    model_out1 = xo1
         
    xo2 = layers.Dense(4096, activation="relu", use_bias=True)(xc)
    xo2 = layers.Dense(1024, activation="relu", use_bias=True)(xo2)
    xo2 = layers.Dense(512, activation="relu", use_bias=True)(xo2)
    xo2 = layers.Dense(1,activation='linear',name="output_2")(xo2)
    model_out2 = xo2


if False:
    """ v8 """
    I5 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_5')
    
    x = _conv_block(I5, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 0}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_109')(x) 
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 16}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_110')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 23}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_111')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 47}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_147')(x) 
    x5 = _conv_block(x, [{'filter': 24, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 3}])
    
    I6 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_6')
    
    x = _conv_block(I6, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 4}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_112')(x)  
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 17}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_113')(x)  
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 22}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_114')(x)  
    x = _conv_block(x,  [{'filter': 64, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 42}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_144')(x) 
    x6 = _conv_block(x, [{'filter': 24, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 7}])
    
    I7 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_7')
    
    x = _conv_block(I7, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 8}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_115')(x) 
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 18}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_116')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 26}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_117')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 45}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_145')(x) 
    x7 = _conv_block(x, [{'filter': 24, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 11}])
    
    I8 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_8')
    
    x = _conv_block(I8, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 12}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_118')(x) 
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 19}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_119')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 20}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_120')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 46}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_146')(x) 
    x8 = _conv_block(x, [{'filter': 24, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 15}])
    
    I1 = Input(shape=(400,), name="input_1")
    x = layers.Dense(400, use_bias=True)(I1)
    x = LeakyReLU(alpha=0.1)(x)
    x = layers.Dense(600, use_bias=True)(x)
    x1 = LeakyReLU(alpha=0.1)(x)
    
    I2 = Input(shape=(400,), name='input_2')
    x = layers.Dense(400, activation="relu", use_bias=True)(I2)
    x = LeakyReLU(alpha=0.1)(x)
    x = layers.Dense(600, activation="relu", use_bias=True)(x)
    x2 = LeakyReLU(alpha=0.1)(x)
    
    I3 = Input(shape=(400,), name='input_3')
    x = layers.Dense(400, activation="relu", use_bias=True)(I3)
    x = LeakyReLU(alpha=0.1)(x)
    x = layers.Dense(600, activation="relu", use_bias=True)(x)
    x3 = LeakyReLU(alpha=0.1)(x)
    
    I4 = Input(shape=(400,), name='input_4')
    x = layers.Dense(400, activation="relu", use_bias=True)(I4)
    x = LeakyReLU(alpha=0.1)(x)
    x = layers.Dense(600, activation="relu", use_bias=True)(x)
    x4 = LeakyReLU(alpha=0.1)(x)
    
    
    xc5=layers.Flatten()(x5)
    xc6=layers.Flatten()(x6)
    xc7=layers.Flatten()(x7)
    xc8=layers.Flatten()(x8)
    
    xc = concatenate([x1,x2,x3,x4,xc5,xc6,xc7,xc8],  name='concat_1')
    xc = layers.Dense(8192, use_bias=True)(xc)
    xc = LeakyReLU(alpha=0.1)(xc)
    
    xo1 = layers.Dense(4096, use_bias=True)(xc)
    xo1 = LeakyReLU(alpha=0.1)(xo1)
    xo1 = layers.Dense(4096, use_bias=True)(xo1)
    xo1 = LeakyReLU(alpha=0.1)(xo1)
    xo1 = layers.Dense(1024, use_bias=True)(xo1)
    xo1 = LeakyReLU(alpha=0.1)(xo1)
    xo1 = layers.Dense(512, use_bias=True)(xo1)
    xo1 = LeakyReLU(alpha=0.1)(xo1)
    xo1 = layers.Dense(1,activation='linear',name="output_1")(xo1)
    model_out1 = xo1
         
    xo2 = layers.Dense(4096, use_bias=True)(xc)
    xo2 = LeakyReLU(alpha=0.1)(xo2)
    xo2 = layers.Dense(1024, use_bias=True)(xo2)
    xo2 = LeakyReLU(alpha=0.1)(xo2)
    xo2 = layers.Dense(512, use_bias=True)(xo2)
    xo2 = LeakyReLU(alpha=0.1)(xo2)
    xo2 = layers.Dense(1,activation='linear',name="output_2")(xo2)
    model_out2 = xo2

if True:
    """ v8 """
    I5 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_5')
    
    x = _conv_block(I5, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 0}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_109')(x) 
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 16}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_110')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 23}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_111')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 47}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_147')(x) 
    x5 = _conv_block(x, [{'filter': 64, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 3}])
    x5 = layers.GlobalAveragePooling2D()(x5)
    
    I6 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_6')
    
    x = _conv_block(I6, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 4}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_112')(x)  
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 17}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_113')(x)  
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 22}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_114')(x)  
    x = _conv_block(x,  [{'filter': 64, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 42}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_144')(x) 
    x6 = _conv_block(x, [{'filter': 64, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 7}])
    x6 = layers.GlobalAveragePooling2D()(x6)
    
    I7 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_7')
    
    x = _conv_block(I7, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 8}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_115')(x) 
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 18}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_116')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 26}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_117')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 45}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_145')(x) 
    x7 = _conv_block(x, [{'filter': 64, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 11}])
    x7 = layers.GlobalAveragePooling2D()(x7)
    
    I8 = Input(shape=(NETWORK_H, NETWORK_W, 1), name='input_8')
    
    x = _conv_block(I8, [{'filter': 6, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 12}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_118')(x) 
    x = _conv_block(x,  [{'filter': 12, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 19}])
    x = MaxPool2D(pool_size=(2, 2), strides=(2, 2), name = 'layer_119')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 20}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_120')(x) 
    x = _conv_block(x,  [{'filter': 32, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 46}])
    x = MaxPool2D(pool_size=(3, 3), strides=(3, 3), name = 'layer_146')(x) 
    x8 = _conv_block(x, [{'filter': 64, 'kernel': 3, 'stride': 1, 'bnorm': False, 'activ': 1, 'layer_idx': 15}])
    x8 = layers.GlobalAveragePooling2D()(x8)
    
    I1 = Input(shape=(400,1), name="input_1")
    x = layers.Conv1D(32,4, use_bias=True)(I1)
    x = LeakyReLU(alpha=0.1)(x)
    x = layers.MaxPooling1D(2)(x)
    x = layers.Conv1D(64,4, use_bias=True)(x)
    x = LeakyReLU(alpha=0.1)(x)
    x1 = layers.MaxPooling1D(2)(x)
    x1 = layers.LSTM(64,return_sequences=False)(x1)
    
    
    I2 = Input(shape=(400,1), name='input_2')
    x = layers.Conv1D(32,4, use_bias=True)(I2)
    x = LeakyReLU(alpha=0.1)(x)
    x = layers.MaxPooling1D(2)(x)
    x = layers.Conv1D(64,4, use_bias=True)(x)
    x = LeakyReLU(alpha=0.1)(x)
    x2 = layers.MaxPooling1D(2)(x)
    x2 = layers.LSTM(64,return_sequences=False)(x2)
    
    
    I3 = Input(shape=(400,1), name='input_3')
    x = layers.Conv1D(32,4, use_bias=True)(I3)
    x = LeakyReLU(alpha=0.1)(x)
    x = layers.MaxPooling1D(2)(x)
    x = layers.Conv1D(64,4, use_bias=True)(x)
    x = LeakyReLU(alpha=0.1)(x)
    x3 = layers.MaxPooling1D(2)(x)
    x3 = layers.LSTM(64,return_sequences=False)(x3)
    
    
    I4 = Input(shape=(400,1), name='input_4')
    x = layers.Conv1D(32,4, use_bias=True)(I4)
    x = LeakyReLU(alpha=0.1)(x)
    x = layers.MaxPooling1D(2)(x)
    x = layers.Conv1D(64,4, use_bias=True)(x)
    x = LeakyReLU(alpha=0.1)(x)
    x4 = layers.MaxPooling1D(2)(x)
    x4 = layers.LSTM(64,return_sequences=False)(x4)
    
    
    
    xc5=layers.Flatten()(x5)
    xc6=layers.Flatten()(x6)
    xc7=layers.Flatten()(x7)
    xc8=layers.Flatten()(x8)
    
    xc = concatenate([x1,x2,x3,x4,xc5,xc6,xc7,xc8],  name='concat_1')
    xc = layers.Dense(1024, use_bias=True)(xc)
    xc = LeakyReLU(alpha=0.1)(xc)
    
    xo1 = layers.Dense(512, use_bias=True)(xc)
    xo1 = LeakyReLU(alpha=0.1)(xo1)
    xo1 = layers.Dense(128, use_bias=True)(xo1)
    xo1 = LeakyReLU(alpha=0.1)(xo1)
    xo1 = layers.Dense(1,activation='linear',name="output_1")(xo1)
    model_out1 = xo1
         
    xo2 = layers.Dense(512, use_bias=True)(xc)
    xo2 = LeakyReLU(alpha=0.1)(xo2)
    xo2 = layers.Dense(128, use_bias=True)(xo2)
    xo2 = LeakyReLU(alpha=0.1)(xo2)
    xo2 = layers.Dense(1,activation='linear',name="output_2")(xo2)
    model_out2 = xo2
   
###################################################################################################################################################

model = Model(inputs=[I1,I2,I3,I4,I5,I6,I7,I8],  outputs={"output_1": model_out1,"output_2": model_out2})
model.summary()




grouping= False
epoch=100


data_generator = Data_Generator(
    batch_size=8,
    root_path="D:\\Projekty\\2022_01_BattPor\\DATA_M1_v5\\",
    IMG_SIZE=(332, 332),
    type_gen="Train",                   # Use trainign data
    split_ratio=0.8,
    input_dirs=["ch4_input_data"],            # adjust to your folder(s)
    label_dirs=["ch4_output_data"],
    use_cached_extrems=True,
    run_extrem=True                   # Do not recalculate if you're just predicting
)

data_train_x=data_generator.data_x
data_train_y=data_generator.data_y
extrems_train=data_generator.extrem_values
intput1, output1= data_generator.__getitem__(1)
all_data=data_generator.all_data

model.compile(optimizer= 'adam', loss={"output_1":losses.MeanSquaredError(),"output_2":losses.MeanSquaredError()},
          loss_weights={"output_1": 1,"output_2": 1})

"""name="model_modul1_v8_best"
model = load_model("T:/K/Klarak-Jaromir/"+name+".h5")"""
earlyStopping = EarlyStopping(monitor='loss', patience=50, verbose=0, mode='min')
mcp_save = ModelCheckpoint(name+'_best.h5', save_best_only=True, monitor='loss', mode='min')



history=model.fit(data_generator,epochs=epoch,batch_size=32,callbacks=[earlyStopping, mcp_save]) 
hist=history.history

model.save("D:/Projekty/2022_01_BattPor/DATA_M1_v5/"+name+".h5") 



data_generator = Data_Generator(
    batch_size=8,
    root_path="D:\\Projekty\\2022_01_BattPor\\DATA_M1_v5\\",
    IMG_SIZE=(332, 332),
    type_gen="Train",                   # Use trainign data
    split_ratio=0.8,
    input_dirs=["ch4_input_data"],            # adjust to your folder(s)
    label_dirs=["ch4_output_data"],
    use_cached_extrems=True,
    run_extrem=True                   # Do not recalculate if you're just predicting
    )

model.compile(optimizer= 'adam', loss={"output_1":losses.MeanSquaredError(),"output_2":losses.MeanSquaredError()},
          loss_weights={"output_1": 1,"output_2": 1})

"""name="model_modul1_v8_best"
model = load_model("T:/K/Klarak-Jaromir/"+name+".h5")"""
earlyStopping = EarlyStopping(monitor='loss', patience=50, verbose=0, mode='min')
#mcp_save = ModelCheckpoint(name+'_best.h5', save_best_only=True, monitor='loss', mode='min')

best_res=5
histor=list()
for epoch in range(150): 
    data_generator.type_gen="Train"
    data_train_x, data_train_y = data_generator.init_data()
    #history=model.fit(data_generator,epochs=epoch,batch_size=32,callbacks=[earlyStopping, mcp_save]) 
    history=model.fit(data_generator,epochs=1,batch_size=32,callbacks=[earlyStopping],verbose=0) 
    hist=history.history["output_1_loss"]
    histor.append(hist[0])
    print(f"Epoch: {epoch} \t loss: {np.round(histor[-1],2)}")
    if float(hist[0])<10:
        data_generator.type_gen="Test"
        data_train_x, data_train_y = data_generator.init_data()
        
        results=np.zeros((len(data_train_x),6),dtype="float32")
        data_train_x=data_generator.data_x
        data_train_y=data_generator.data_y
        list_of_keys=list()
        for idx in range(int(len(data_train_x)//data_generator.batch_size)):
            out_test_0 =np.zeros((len(data_train_x),6),dtype="float32")
            trainx_idx, trainy_idx=data_generator.__getitem__(idx)
            data_v1=trainx_idx["input_1"]
            data_v2=trainx_idx["input_2"]
            res=model.predict(trainx_idx,verbose=0)
            for i in range(data_generator.batch_size):
                results[idx*data_generator.batch_size+i,0]=np.round(np.float32(trainy_idx[list(res.keys())[0]][i]),2)
                results[idx*data_generator.batch_size+i,1]=np.round(np.float32((res[list(res.keys())[0]])[i][0]),2)
                
                results[idx*data_generator.batch_size+i,3]=np.round(np.float32(trainy_idx[list(res.keys())[1]][i]),2)
                results[idx*data_generator.batch_size+i,4]=np.round(np.float32((res[list(res.keys())[1]])[i][0]),2)
                list_of_keys.append(data_train_x[idx*data_generator.batch_size+i])
          
        results[:,2]=results[:,0]-results[:,1]
                    
        results[:,5]=results[:,3]-results[:,4]
        print(f"Epoch: {epoch} \tRESULT: {np.round(float(np.mean(np.abs(results[:,2]))),2)} % and loss: {np.round(float(histor[-1]),2)}")
        if best_res>np.mean(np.abs(results[:,2])):
            print(f"Storing best model in epoch: {epoch} result: {np.round(np.mean(np.abs(results[:,2])),2)} % and lost: {histor[-1]}")
            best_res=np.mean(np.abs(results[:,2]))
            model.save("D:/Projekty/2022_01_BattPor/DATA_M1_v5/"+name+"_"+str(epoch)+".h5") 
        

pd_output=pd.DataFrame(columns=["Name","True porosity","Predicted porosity", "Error in porosity", "True thickness","Predicted thickness", "Error in thickness"])
pd_output["Name"]=list_of_keys
pd_output["True porosity"]=np.round(results[:,0],2)
pd_output["Predicted porosity"]=np.round(results[:,1],3)
pd_output["Error in porosity"]=np.round(results[:,2],3)
pd_output["True thickness"]=np.round(results[:,3],2)
pd_output["Predicted thickness"]=np.round(results[:,4],3)
pd_output["Error in thickness"]=np.round(results[:,5],3)

sorted_i=np.argsort(np.asarray(list_of_keys))
sorted_df = pd_output.reindex(sorted_i).reset_index(drop=True)  
sorted_df.to_excel("Training results.xlsx")

loaded_excel=openpyxl.load_workbook("Training results.xlsx")
wb=loaded_excel.active
wb["K2"]="Mean error of porosity:"
wb["L2"]=np.mean(np.abs(results[:,2]))
wb["K3"]="Standard deviation of porosity error:"  
wb["L3"]=np.std(np.abs(results[:,2]))
wb["K4"]="Mean error of thickness:"
wb["L4"]=np.mean(np.abs(results[:,5]))
wb["K5"]="Standard deviation of thickness error:"  
wb["L5"]=np.std(np.abs(results[:,5]))
bold_font = openpyxl.styles.Font(bold=True)
yel = openpyxl.styles.PatternFill(start_color="FFFF00",fill_type="solid")

for row in wb.iter_rows(min_row=2, min_col=11, max_row=5, max_col=12):
    for cell in row:
        cell.fill=yel
        cell.font=bold_font

loaded_excel.save("Training results_"+name+".xlsx")

    
pd_extrems=pd.DataFrame({"Min":data_generator.extrem_values[:,0],"Max":data_generator.extrem_values[:,1]})
pd_extrems.to_json("extrems.json")



data_generator.type_gen="Test"
data_x,data_y =  data_generator.get_data()
data_train_x=data_x
data_train_y=data_y
results=np.zeros((len(data_train_x),6),dtype="float32")
list_of_keys=list()
for idx in range(int(len(data_train_x)//data_generator.batch_size)):
    out_test_0 =np.zeros((len(data_train_x),6),dtype="float32")
    trainx_idx, trainy_idx=data_generator.__getitem__(idx)
    data_v1=trainx_idx["input_1"]
    data_v2=trainx_idx["input_2"]
    res=model.predict(trainx_idx)
    for i in range(data_generator.batch_size):
        results[idx*data_generator.batch_size+i,0]=np.round(np.float32(trainy_idx[list(res.keys())[0]][i]),2)
        results[idx*data_generator.batch_size+i,1]=np.round(np.float32((res[list(res.keys())[0]])[i][0]),2)
        
        results[idx*data_generator.batch_size+i,3]=np.round(np.float32(trainy_idx[list(res.keys())[1]][i]),2)
        results[idx*data_generator.batch_size+i,4]=np.round(np.float32((res[list(res.keys())[1]])[i][0]),2)
        list_of_keys.append(data_train_x[idx*data_generator.batch_size+i])
  
results[:,2]=results[:,0]-results[:,1]
print(np.mean(np.abs(results[:,2])))
print(np.std(np.abs(results[:,2])))
            
results[:,5]=results[:,3]-results[:,4]
print(np.mean(np.abs(results[:,5])))
print(np.std(np.abs(results[:,5])))
        
    

pd_output=pd.DataFrame(columns=["Name","True porosity","Predicted porosity", "Error in porosity", "True thickness","Predicted thickness", "Error in thickness"])
pd_output["Name"]=list_of_keys
pd_output["True porosity"]=np.round(results[:len(list_of_keys),0],2)
pd_output["Predicted porosity"]=np.round(results[:len(list_of_keys),1],3)
pd_output["Error in porosity"]=np.round(results[:len(list_of_keys),2],3)
pd_output["True thickness"]=np.round(results[:len(list_of_keys),3],2)
pd_output["Predicted thickness"]=np.round(results[:len(list_of_keys),4],3)
pd_output["Error in thickness"]=np.round(results[:len(list_of_keys),5],3)

sorted_i=np.argsort(np.asarray(list_of_keys))
key_srtoed=list_of_keys[sorted_i]
pd_output.index=sorted_i
sorted_df = pd_output.sort_index(ascending=True)  
sorted_df.to_excel("Training results.xlsx")

loaded_excel=openpyxl.load_workbook("Training results.xlsx")
wb=loaded_excel.active
wb["K2"]="Mean error of porosity:"
wb["L2"]=np.mean(np.abs(results[:,2]))
wb["K3"]="Standard deviation of porosity error:"  
wb["L3"]=np.std(np.abs(results[:,2]))
wb["K4"]="Mean error of thickness:"
wb["L4"]=np.mean(np.abs(results[:,5]))
wb["K5"]="Standard deviation of thickness error:"  
wb["L5"]=np.std(np.abs(results[:,5]))
bold_font = openpyxl.styles.Font(bold=True)
yel = openpyxl.styles.PatternFill(start_color="FFFF00",fill_type="solid")

for row in wb.iter_rows(min_row=2, min_col=11, max_row=5, max_col=12):
    for cell in row:
        cell.fill=yel
        cell.font=bold_font

loaded_excel.save("Training results_intest_"+name+".xlsx")







indexes = [1, 2, 3, 4, 7, 8, 9, 10]
train_x = []
train_y = []
train_x_pre = np.zeros((data_generator.batch_size, 8), dtype="object")
stop = np.zeros((data_generator.batch_size, 2), dtype="float32")
train_y_pre = np.zeros((data_generator.batch_size, 2), dtype="float32")

for cdx in range(data_generator.batch_size):
    itemx_path = os.path.join(data_generator.data_x[data_generator.batch_size * idx + cdx])
    itemy_path = os.path.join(data_generator.data_y[data_generator.batch_size * idx + cdx])

    datax_idx = pd.read_json(itemx_path)
    with open(itemy_path, "r") as f:
        datay_idx = np.fromstring(f.read(), dtype=np.float32, sep=' ')
    datax_idx_val = datax_idx.values
    strh = int(datax_idx_val[0, 5])
    stop[cdx, 0] = strh

    for indx, ind in enumerate(indexes):
        train_x_pre[cdx, indx] = datax_idx_val[0, ind]
    train_y_pre[cdx, :] = datay_idx

for indx, ind in enumerate(indexes):
    if indx < 4:
        train_x_post = np.zeros((data_generator.batch_size, 400), dtype="float")
        for idxx in range(data_generator.batch_size):
            norm_data = train_x_pre[idxx, indx][int(stop[idxx, 0])-20:int(stop[idxx, 0])+380]
            train_x_post[idxx, :] = norm_data / data_generator.extrem_values[indx, 1]
    else:
        train_x_post = np.zeros((data_generator.batch_size, data_generator.img_size[0], data_generator.img_size[1]), dtype="float")
        for idxx in range(data_generator.batch_size):
            train_x_post[idxx, :, :] = train_x_pre[idxx, indx] / data_generator.extrem_values[indx, 1]
    train_x.append(train_x_post)

train_y = np.asarray([train_y_pre[:, 0], train_y_pre[:, 1]])













